from pathlib import Path

import openpyxl

from watcher.converters.base import ConversionResult


def convert(path: Path) -> ConversionResult:
    """Reads every sheet and renders it as a markdown table — deterministic,
    no LLM call. Cell values only; formulas aren't evaluated beyond
    whatever openpyxl's data_only mode already resolved on last save."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(cell.strip() for cell in r)]  # drop fully-empty rows
        if not rows:
            continue
        sections.append(f"## {sheet_name}\n\n{_rows_to_markdown_table(rows)}")
    wb.close()

    text_content = f"# {path.stem}\n\n" + "\n\n".join(sections) if sections else f"# {path.stem}\n\n(empty workbook)"
    return ConversionResult(text_content=text_content, suggested_okf_type="Worksheet")


def _rows_to_markdown_table(rows: list) -> str:
    header, *body = rows
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    lines += ["| " + " | ".join(r) + " |" for r in body]
    return "\n".join(lines)
